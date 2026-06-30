#!/usr/bin/env python3

import argparse
import csv
import json
import re
from pathlib import Path


HEADERS = [
    "用例ID",
    "原始候选ID",
    "优先级",
    "评审状态",
    "用例标题",
    "需求追溯",
    "测试点追溯",
    "前置条件",
    "测试步骤",
    "预期结果",
    "证据要求",
    "标签",
]


def main():
    parser = argparse.ArgumentParser(description="Export Harness accepted-test-cases.yaml to Markdown, CSV, and XLSX.")
    parser.add_argument("--input", required=True, help="Path to accepted-test-cases.yaml")
    parser.add_argument("--out-dir", required=True, help="Output directory")
    args = parser.parse_args()

    input_path = Path(args.input)
    out_dir = Path(args.out_dir)
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")
    out_dir.mkdir(parents=True, exist_ok=True)

    cases = parse_cases(input_path.read_text(encoding="utf-8"))
    write_csv(out_dir / "test-cases.csv", cases)
    write_markdown(out_dir / "test-cases.md", cases)
    xlsx_path = write_xlsx(out_dir / "test-cases.xlsx", cases)

    summary = {
        "total": len(cases),
        "p0": sum(1 for item in cases if item["优先级"] == "P0"),
        "p1": sum(1 for item in cases if item["优先级"] == "P1"),
        "p2": sum(1 for item in cases if item["优先级"] == "P2"),
        "paths": {
            "markdown": str(out_dir / "test-cases.md"),
            "csv": str(out_dir / "test-cases.csv"),
            "xlsx": str(xlsx_path) if xlsx_path else "",
        },
    }
    (out_dir / "test-cases-export-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def parse_cases(text):
    cases = []
    for part in text.split("\n  - id: ")[1:]:
        block = "    id: " + part
        item = {
            "用例ID": first(block, "id"),
            "原始候选ID": first(block, "original_candidate_id"),
            "优先级": nested_first(block, "risk", "level"),
            "评审状态": nested_first(block, "review", "status"),
            "用例标题": first(block, "title"),
            "需求追溯": ", ".join(nested_list(block, "traceability", "requirement_ids")),
            "测试点追溯": ", ".join(nested_list(block, "traceability", "function_point_ids")),
            "前置条件": "\n".join(top_list(block, "preconditions")),
            "测试步骤": "\n".join(step_lines(block)),
            "预期结果": nested_first(block, "expected_result", "summary"),
            "证据要求": ", ".join(nested_list(block, "expected_result", "evidence")),
            "标签": ", ".join(top_list(block, "tags")),
        }
        if item["用例ID"]:
            cases.append(item)
    return cases


def clean(value):
    value = (value or "").strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in ["'", '"']:
        value = value[1:-1]
    return value.replace('\\"', '"')


def first(block, key):
    match = re.search(rf"^    {re.escape(key)}:\s*(.*)$", block, re.M)
    return clean(match.group(1)) if match else ""


def nested_block(block, parent):
    match = re.search(rf"^    {re.escape(parent)}:\n([\s\S]*?)(?=\n    \S|\n  - id:|\Z)", block, re.M)
    return match.group(1) if match else ""


def nested_first(block, parent, key):
    section = nested_block(block, parent)
    match = re.search(rf"^      {re.escape(key)}:\s*(.*)$", section, re.M)
    return clean(match.group(1)) if match else ""


def nested_list(block, parent, key):
    section = nested_block(block, parent)
    match = re.search(rf"^      {re.escape(key)}:\n((?:        - .*\n?)+)", section, re.M)
    if not match:
        return []
    return [clean(item) for item in re.findall(r"^        -\s*(.*)$", match.group(1), re.M)]


def top_list(block, key):
    match = re.search(rf"^    {re.escape(key)}:\n((?:      - .*\n?)+)", block, re.M)
    if not match:
        return []
    return [clean(item) for item in re.findall(r"^      -\s*(.*)$", match.group(1), re.M)]


def step_lines(block):
    match = re.search(r"^    steps:\n([\s\S]*?)(?=\n    validators:|\n    expected_result:|\n    \S|\n  - id:|\Z)", block, re.M)
    if not match:
        return []
    raw = match.group(1)
    actions = [clean(item) for item in re.findall(r"^        action:\s*(.*)$", raw, re.M)]
    expects = [clean(item) for item in re.findall(r"^        expected_observation:\s*(.*)$", raw, re.M)]
    lines = []
    for index, action in enumerate(actions, 1):
        expected = expects[index - 1] if index - 1 < len(expects) else ""
        lines.append(f"{index}. {action}" + (f" -> {expected}" if expected else ""))
    return lines


def write_csv(path, cases):
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(cases)


def write_markdown(path, cases):
    lines = [
        "# Harness 测试用例",
        "",
        f"- 用例总数: {len(cases)}",
        f"- P0: {sum(1 for item in cases if item['优先级'] == 'P0')}",
        f"- P1: {sum(1 for item in cases if item['优先级'] == 'P1')}",
        f"- P2: {sum(1 for item in cases if item['优先级'] == 'P2')}",
        "",
        "## 用例清单",
        "",
    ]
    for item in cases:
        lines.extend([
            f"### {item['用例ID']} {item['用例标题']}",
            "",
            f"- 优先级: {item['优先级']}",
            f"- 评审状态: {item['评审状态']}",
            f"- 追溯: {item['需求追溯']}; {item['测试点追溯']}",
            f"- 前置条件: {item['前置条件']}",
            "",
            "测试步骤:",
            "",
            item["测试步骤"] or "N/A",
            "",
            f"预期结果: {item['预期结果']}",
            "",
        ])
    path.write_text("\n".join(lines), encoding="utf-8")


def write_xlsx(path, cases):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试用例"
    sheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in cases:
        sheet.append([item[header] for header in HEADERS])

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [16, 16, 10, 14, 42, 26, 22, 32, 70, 56, 24, 28]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)
    return path


if __name__ == "__main__":
    main()
